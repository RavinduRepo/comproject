# import required libraries
import openpyxl
import time
from nltk.corpus import wordnet
import random
import turtle
from tkinter import *
from pyfirmata import Arduino, util, INPUT, OUTPUT

# set arduino board and set pins according to LDR, LED and BUTTONS
board = Arduino("COM4")
led_pin = 10
led_pin_1 = 12
led_pin_2 = 13
led_pin_3 = 1
led1, led2, led3, led4, led5, led6, led7, led8, led9 = 2, 3, 4, 5, 6, 7, 8, 9, 10
board.digital[led1].mode = OUTPUT
board.digital[led2].mode = OUTPUT
board.digital[led3].mode = OUTPUT
board.digital[led4].mode = OUTPUT
board.digital[led5].mode = OUTPUT
board.digital[led6].mode = OUTPUT
board.digital[led7].mode = OUTPUT
board.digital[led8].mode = OUTPUT
board.digital[led9].mode = OUTPUT
board.digital[led_pin_1].mode = INPUT
board.digital[led_pin_2].mode = INPUT
board.analog[led_pin_3].mode = INPUT

board.digital[led_pin].mode = OUTPUT
ldr_pin = 0
board.analog[ldr_pin].mode = INPUT
it = util.Iterator(board)
it.start()
board.analog[ldr_pin].enable_reporting()
board.analog[led_pin_3].enable_reporting()
lo2 = openpyxl.load_workbook('ACC.xlsx')  # open exel sheet to write and read. this is for create new account and get
# the rank
sheet = lo2.sheetnames
sht1 = lo2[sheet[0]]
word_file = open('word.txt', 'r')  # open the .txt file for reading
word_list = word_file.readlines()  # word_list is the list containing words for the game
used_list, quic, quic_0 = [], [], []
dif_lev, ww, show = 1, [], 'EASY'
min_1 = 0
max_1 = 0
row_1 = 0
case = 'no'
word_1, word_z = '', ''
dict_1 = 'none'
attempt = 3
status = 'none'
c_unlock = []
im = 1
m_list = ['01', '1000', '1010', '100', '0', '0010', '110', '0000', '00', '0111', '101', '0100', '11', '10', '111',
          '0110', '1101', '010', '000', '1', '001', '0001', '011', '1001', '1011', '1100']
let1 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O'
    , 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
dl_3_l, dl_2_l, dl_1_l = [], [], []
for i in word_list:  # this for loop categorized words according to length of the word
    if len(i[:-1]) > 8:
        dl_3_l.append(i[:])
    if 6 < len(i[:-1]) < 9:
        dl_2_l.append(i[:])
    if 4 < len(i[:-1]) < 8:
        dl_1_l.append(i[:])
root = Tk()


def get_acc_name(event=None):  # this function run when user press 'enter' key after entering his account name
    acc = get_acc.get().upper()
    get_acc.delete(0, 'end')
    if len(acc) == 0:
        return False
    t0 = 1
    while True:
        p = sht1.cell(row=t0, column=1).value
        if p == acc:
            notice.delete(0, 'end')
            notice.insert(END, " " * 10 + 'WELLCOME ' + str(acc))
            global used_list
            used_list = list(sht1.cell(row=t0, column=6).value)
            global row_1
            row_1 = t0
            break
        if p is None:  # if account name is new one this will update the exel sheet
            notice.delete(0, 'end')
            notice.insert(END, " " * 10 + 'WELLCOME NEW PLAYER')
            sht1.cell(row=t0, column=1).value = acc
            sht1.cell(row=t0, column=2).value = 0
            sht1.cell(row=t0, column=3).value = 0
            sht1.cell(row=t0, column=4).value = 0
            sht1.cell(row=t0, column=5).value = 1
            sht1.cell(row=t0, column=6).value = '0'
            used_list = ['0']
            lo2.save('ACC.xlsx')
            break
        t0 += 1

def exit_board():
    global board
    board.exit()

def show_details():  # this function runs after press the 'ranking' button
    notice.delete(0, 'end')
    t = 1
    list_1, list_2, list_3, list_4, list_5 = [], [], [], [], []
    while True:
        p1 = sht1.cell(row=t, column=1).value
        if not (p1 is None):  # get the values from the exel sheet and appends to the lists
            list_1.append(sht1.cell(row=t, column=1).value)
            list_2.append(sht1.cell(row=t, column=2).value)
            list_3.append(sht1.cell(row=t, column=3).value)
            list_4.append(sht1.cell(row=t, column=4).value)
            list_5.append(sht1.cell(row=t, column=5).value)
            t += 1
        else:
            break
    points = []
    for i in range(len(list_1)):
        points.append((int(list_2[i]) + int(list_3[i])) + int(list_4[i]) / list_5[i])  # calculate the points
    for i in range(len(list_1)):  # show profile details according to points
        p = points.index(max(points))
        string_1 = ' ' + str(list_1[p]) + ' >>> RANK = ' + str(i + 1) + '  easy = ' + str(
            list_2[p]) + '  hard = ' + str(list_3[p]) + '  points = ' + str(points[p])
        points[p] = -1
        notice.insert(END, string_1)
        notice.insert(END, "  ")


# following draw functions draw specific part of the full hang man image
def draw_1():
    pen.speed(10)
    pen.up()
    pen.left(90)
    pen.forward(150)
    pen.left(90)
    pen.forward(100)
    pen.down()
    pen.speed(5)
    pen.width(5)
    pen.back(150)
    pen.forward(50)
    pen.left(90)


def draw_2():
    pen.forward(50)


def draw_3():
    pen.up()
    pen.forward(30)
    pen.right(90)
    pen.forward(30)
    pen.left(90)
    pen.down()
    pen.circle(30)


def draw_4():
    pen.up()
    pen.forward(30)
    pen.left(90)
    pen.forward(30)
    pen.right(90)
    pen.down()
    pen.forward(12)


def draw_5():
    pen.right(45)
    pen.forward(60)
    pen.back(60)


def draw_6():
    pen.left(90)
    pen.forward(60)
    pen.back(60)


def draw_7():
    pen.right(45)
    pen.forward(120)


def draw_8():
    pen.right(45)
    pen.forward(60)
    pen.back(60)


def draw_9():
    pen.left(90)
    pen.forward(60)
    pen.back(60)
    pen.right(45)


def main_draw():  # this function draw the hang man image using above functions. according to the difficulty level
    # this function draw the hang man in 3 or 6 or 9 steps
    global dif_lev, attempt
    if dif_lev == 1:
        if attempt == 7:
            draw_1()
            draw_2()
        elif attempt == 6:
            draw_3()
        elif attempt == 5:
            draw_4()
        elif attempt == 4:
            draw_5()
        elif attempt == 3:
            draw_6()
        elif attempt == 2:
            draw_7()
        elif attempt == 1:
            draw_8()
        else:
            draw_9()
    if dif_lev == 2:
        if attempt == 5:
            draw_1()
            draw_2()
        elif attempt == 4:
            draw_3()
            draw_4()
        elif attempt == 3:
            draw_5()
            draw_6()
        elif attempt == 2:
            draw_7()
        elif attempt == 1:
            draw_8()
        else:
            draw_9()
    if dif_lev == 3:
        if attempt == 2:
            draw_1()
            draw_2()
            draw_3()
        elif attempt == 1:
            draw_4()
            draw_5()
            draw_6()
        else:
            draw_7()
            draw_8()
            draw_9()


def main():  # this function run after the user press the play button.
    hang_man.delete('all')
    global defi_level, max_1, min_1, word_1, attempt, dict_1, dl_3_l, dl_2_l, dl_1_l, ww, pen, quic, quic_0, case, c_unlock
    pen = turtle.RawTurtle(hang_man)
    # for a new round all the variables ara rest
    case = 'no'
    quic = []
    quic_0 = []
    c_unlock = []
    for iw in range(9):
        board.digital[iw + 2].write(0)
    word.configure(text='H A N G _ _ _ M A N')
    if len(used_list) == 0:
        notice.delete(0, 'end')
        notice.insert(END, 'PLEASE ENTER YOUR ACCOUNT NAME')
        notice.insert(E, '(if you dont have an account please give a new account name)')
        return False
    if dif_lev == 3:
        for iw in range(3):
            board.digital[iw + 2].write(1)
        max_1 = 12
        min_1 = 7
        attempt = 3
        ww = dl_3_l
    if dif_lev == 2:
        for iw in range(6):
            board.digital[iw + 2].write(1)
        max_1 = 9
        min_1 = 5
        attempt = 6
        ww = dl_2_l
    if dif_lev == 1:
        for iw in range(8):
            board.digital[iw + 2].write(1)
        max_1 = 7
        min_1 = 4
        attempt = 9
        ww = dl_1_l
    notice.delete(0, 'end')
    while True:  # this will choose a word randomly from the word lists
        index = random.randint(0, len(ww) - 1)
        word_1 = ww[index][:-1].upper()
        if str(index) in used_list:
            continue
        if min_1 <= len(word_1) <= max_1:
            used_list.append(str(index))
            break
    notice.delete(0, 'end')
    length = len(word_1)
    print(word_1)
    word.configure(text='_ ' * length)


def check(event=None):  # this function runs when user press 'enter' key after enter his answer in entry box
    global word_1, attempt, quic, quic_0, case, c_unlock
    if len(used_list) == 0:
        notice.delete(0, 'end')
        notice.insert(END, 'PLEASE ENTER YOUR ACCOUNT NAME')
        notice.insert(E, '(if you dont have an account please give a new account name)')
        return False
    if im == 1:
        word_2 = guess.get().upper()  # this will get the users answer if it enters using keyboard
    else:
        word_2 = word_z  # this will get the users answer if it enters using morse codes
    if len(word_2) != len(word_1):  # check weather the real word and answer is in same length or not
        notice.delete(0, 'end')
        notice.insert(END, 'YOUR ANSWER IS INCORRECT')
        attempt -= 1  # change the remaining attempts
        print('==== ', attempt)
        for iw in range(9):
            board.digital[iw + 2].write(0)
        for iw in range(attempt):
            board.digital[iw + 2].write(1)
        if attempt == 0:
            notice.delete(0, 'end')
            notice.insert(END, 'YOU LOSE !!')
            # if attempts is over then resets the variable for new round
            quic = []
            quic_0 = []
            case = 'no'
            word_1 = ''
            c_unlock = []
            word.configure(text='H A N G _ _ _ M A N')
        main_draw()
        guess.delete(0, 'end')
        return False
    if attempt == 0:
        return False
    letter = []
    for i in range(len(word_1)):
        if word_2[i] != word_1[i]:  # check the letters of two words one by one
            notice.delete(0, 'end')
            notice.insert(END, 'YOUR ANSWER ', word_2, ' IS INCORRECT')  # if one of the letter is not same give a
            # notice saying answer is incorrect
        else:
            c_unlock.append(i)
            letter.append(word_2[i])  # if letters are same collect them.
    new_dis = []
    for i in range(len(word_1)):
        if i in c_unlock:
            new_dis.append(' ' + word_1[i])
        else:
            new_dis.append('_ ')
    new_dis_1 = ''.join(new_dis)  # using collected letters makes a string.
    if new_dis.count('_ ') == 0:  # check if the answer is correct.
        notice.delete(0, 'end')
        notice.insert(END, 'YOUR ANSWER IS CORRECT')
        notice.insert(END, 'YOU WIN !!!')
        # if answer is correct rest variables for new round.
        quic = []
        quic_0 = []
        case = 'no'
        hang_man.delete('all')
        word_1 = ''
        c_unlock = []
    else:
        attempt -= 1  # calculate remaining attempts.
        if attempt == 0:  # if attempts is 0 reset variables.
            notice.delete(0, 'end')
            notice.insert(END, 'YOU LOSE !!')
            quic = []
            quic_0 = []
            case = 'no'
            word.configure(text='H A N G _ _ _ M A N')
            word_1 = ''
            c_unlock = []
        main_draw()  # draw the hang man using main draw function.
    if attempt != 0:
        word.configure(text=new_dis_1)
    guess.delete(0, 'end')


def difficulty():  # this function will run when user press difficulty button.
    global status
    status = 'DDD'  # change the status to DDD and change label_1 to label_3 labels texts and colors.
    label_1.configure(text='EASY', bg='green')
    label_2.configure(text='MEDIUM', bg='yellow')
    label_3.configure(text='HARD', bg='red')


def helps_1():  # this function will run when user press help button.
    global status, dif_lev
    status = 'HHH'  # it changes the status to 'HHH' and change label_1 to label_3 labels texts and colors according
    # to current difficulty level.
    if dif_lev == 3:
        label_1.configure(text='DEFINITION', bg='red')
        label_2.configure(text='BUTTON B DESCRIPTION', bg='pink')
        label_3.configure(text='BUTTON C DESCRIPTION', bg='pink')
    if dif_lev == 2:
        label_1.configure(text='DEFINITION', bg='red')
        label_2.configure(text='SHOW SIMILAR WORDS', bg='yellow')
        label_3.configure(text='BUTTON C DESCRIPTION', bg='pink')
    if dif_lev == 1:
        label_1.configure(text='DEFINITION', bg='red')
        label_2.configure(text='SHOW SIMILAR WORDS', bg='yellow')
        label_3.configure(text='SHOW SOME LETTERS', bg='green1')


def input_1():  # this function will run when user press input method button.
    global status
    status = 'III'  # change the status to III and change label_1 to label_3 labels texts and colors.
    label_1.configure(text='KEY BOARD', bg='gray40')
    label_2.configure(text='MORSE CODE', bg='gray40')
    label_3.configure(text='BUTTON C DESCRIPTION', bg='pink')


def AAAA():  # this function will run when user press A button.
    # this function has multiple tasks. this will eliminate them according to value of 'status' variable.
    global status, dif_lev, dict_1, word_1, attempt, case, c_unlock, im
    if status == 'DDD':  # if status = DDD change difficulty level and change total attempts.
        dif_lev = 1
        attempt = 8
        for iw in range(9):
            board.digital[iw + 2].write(0)
        for iw in range(attempt):
            board.digital[iw + 2].write(1)
        cDif.configure(text='EASY')  # change cDif label to  easy
    if status == 'HHH':  # if status  = HHH this will find a definition to the word using nltk module.
        if len(word_1) > 1:
            if len(quic) == 0:
                try:
                    dict_1_1 = wordnet.synsets(word_1)
                    dict_1 = dict_1_1[0].definition()
                    print(dict_1)
                    quic.append(dict_1)
                except:
                    quic.append('')
            else:
                dict_1 = quic[0]
            if dict_1 is None:  # if there is no definition unlock some letters.
                notice.insert(END, 'no definition in PYDICTIONARY')
                if case == 'no':  # before unlock letters check case variable. if case != no then
                    # this will not unlock letters. because the case variable work as an indicator weather user
                    # has unlocked letters or not.
                    notice.insert(END, "BUT WE'LL UNLOCK SOME LETTERS !!")
                    num_l = round(len(word_1) / 3)
                    print(num_l)
                    new_diss = []
                    for k in range(num_l):
                        ran_n = random.randint(0, len(word_1) - 1)
                        c_unlock.append(ran_n)
                    for k in range(len(word_1)):
                        if k in c_unlock:
                            new_diss.append(word_1[k] + ' ')
                        else:
                            new_diss.append('_ ')
                    print(c_unlock, new_diss)
                    str_2 = ''.join(new_diss)
                    word.configure(text=str_2)
                    case = 'yes'  # after unlock letters change the case variable to yes.
                return False
            notice.delete(0, 'end')
            if len(quic[0]) > 80:
                qqq = (len(quic[0]) // 80) + 1
                for iq in range(qqq):
                    a = iq * 80
                    b = (iq + 1) * 80
                    if iq == qqq:
                        notice.insert(END, quic[0][a:])
                    else:
                        notice.insert(END, quic[0][a:b])
            else:
                notice.insert(END, quic[0])
    if status == "III":  # if status = III then give a message to user to use keyboard.
        notice.delete(0, 'end')
        notice.insert(END, 'enter your word using key board')
        im = 1
    if status == 'other':  # if status = 'other' then run morse_code() function.
        morse_code()


def BBBB():  # this function will run when user press B button.
    # this function has multiple tasks. this will eliminate them according to value of 'status' variable.
    global status, dif_lev, dict_1, word_1, attempt, im
    if status == 'DDD':  # if status = DDD change difficulty level and change total attempts.
        dif_lev = 2
        attempt = 6
        for iw in range(9):
            board.digital[iw + 2].write(0)
        for iw in range(attempt):
            board.digital[iw + 2].write(1)
        cDif.configure(text='MEDIUM')  # change cDif label to  medium
    elif status == 'HHH':  # if status  = HHH this will find a definition to the word using nltk module.
        if len(word_1) > 1:
            if dif_lev < 3:
                if len(quic_0) == 0:
                    similar = []
                    for i3 in wordnet.synsets(word_1):
                        for i4 in i3.lemmas():
                            similar.append(i4.name())
                    quic_0.append(similar)
                else:
                    similar = quic_0[0]
                notice.delete(0, 'end')
                lala = set(similar)
                baba = list(lala)
                for syn in baba:
                    if str(syn).lower() in word_1.lower():
                        baba.remove(syn)
                        continue
                    if word_1.lower() in str(syn).lower():
                        baba.remove(syn)
                        continue
                    if word_1.lower() == str(syn).lower():
                        baba.remove(syn)
                        continue
                    notice.insert(END, '>>>' + syn)
                if len(baba) == 0:
                    notice.insert(END, 'sorry no similar word found')
            else:
                return False
    elif status == 'III':  # if status = III change status to 'other' and change im to 0
        # if im = 1 then get users answer using keyboard and if im = 0 then get answer using morse codes
        status = 'other'
        im = 0
        label_1.configure(text='START', bg='green1')
        label_2.configure(text='BUTTON B DESCRIPTION', bg='gray')
        label_3.configure(text='BUTTON C DESCRIPTION', bg='gray')
    elif status == 'other':
        return False


def CCCC():  # this function will run when user press C button.
    # this function has multiple tasks. this will eliminate them according to value of 'status' variable.
    global status, dif_lev, dict_1, word_1, case, attempt, c_unlock
    if status == 'DDD':  # if status = DDD change difficulty level and change total attempts.
        dif_lev = 3
        attempt = 3
        for iw in range(9):
            board.digital[iw + 2].write(0)
        for iw in range(attempt):
            board.digital[iw + 2].write(1)
        cDif.configure(text='HARD')
    if status == 'HHH':  # then unlock some random letters.
        if dif_lev < 2:
            if case == 'no':
                num_l = round(len(word_1) / 3)
                new_diss = []
                for k in range(num_l):
                    ran_n = random.randint(0, len(word_1) - 1)
                    c_unlock.append(ran_n)
                for k in range(len(word_1)):
                    if k in c_unlock:
                        new_diss.append(word_1[k] + ' ')
                    else:
                        new_diss.append('_ ')
                str_2 = ''.join(new_diss)
                word.configure(text=str_2)
                case = 'yes'
            else:
                notice.delete(0, 'end')
                notice.insert(END, ' you have already unlock some letters')
        else:
            return False
    if status == 'III':
        return False
    if status == 'other':
        return False


def morse_code():  # this function will read morse code and get the user answer as word_z
    global word_1, word_z
    time_1 = []
    time_2 = []
    x = 1
    holder = '1'
    code = []
    tt = time.time()
    t11 = -1
    t22 = -1
    t33 = -1
    while True:
        ldr_val = board.analog[ldr_pin].read()
        t1 = time.time()
        if t1 - tt < 1:
            continue
        if x == 1:
            print('start')
            x = 2
        if t11 > 0 and (t1 - t11) > 0.5:
            t11 = -1
            board.digital[led_pin].write(0)
        if t22 > 0 and (t1 - t22) > 2.5:
            t22 = -1
            board.digital[led_pin].write(0)
        if t33 > 0 and (t1 - t33) > 3.5:
            t33 = -1
            board.digital[led_pin].write(0)
        if board.digital[led_pin_1].read():
            board.digital[led_pin].write(1)
            t11 = time.time()
        if board.digital[led_pin_2].read():
            board.digital[led_pin].write(1)
            t22 = time.time()
        if board.analog[led_pin_3].read() > 0.5:
            board.digital[led_pin].write(1)
            t33 = time.time()
        if holder == '1' and ldr_val > 0.0:
            time_1.append(t1)
            holder = '0'
        if holder == '0' and ldr_val == 0.0:
            time_2.append(t1)
            holder = '1'
            p = len(time_1) - 1
            if (time_2[p] - time_1[p]) > 3:
                code.append('/')
            if 0.8 < time_2[p] - time_1[p] < 3:
                code.append('1')
                print(1)
            if time_2[p] - time_1[p] < 0.8:
                print(0)
                code.append('0')
            if code.count('/') == len(word_1):
                break
        time.sleep(0.001)
    ll = ''.join(code[:-1])
    list_11 = ll.split('/')
    print(list_11)
    word_z1 = []
    for let in list_11:
        if let in m_list:
            zz = m_list.index(let)
            print(zz)
            word_z1.append(let1[zz])
        else:
            notice.delete(0, 'end')
            notice.insert(END, 'wrong code try again')
            return False
    word_z = ''.join(word_z1)
    check()  # after get the word_z then run check() function


# this following codes for GUI
acc_name = Label(root, text='ACCOUNT NAME', font=('bold', 15), width=25, fg='red', borderwidth=5, bg='pink')
acc_name.grid(row=0, column=1)
get_acc = Entry(root, width=18, font=('bold', 16))
get_acc.bind('<Return>', get_acc_name)
get_acc.grid(row=0, column=2)
acc_details = Button(root, text='RANKING', font=('bold', 12), command=show_details)
acc_details.grid(row=0, column=3)
notice = Listbox(root, width=80, font=5, height=16)
notice.grid(row=6)
word = Label(text='H A N G _ _ _ M A N', font=('bold', 60), height=1, bg='yellow', fg='green')
word.grid(row=1, column=0)
hang_man = Canvas(root, bg='black', width=300, height=390)
hang_man.grid(row=6, column=1, pady=20)
pen = turtle.RawTurtle(hang_man)
play_but = Button(root, font=('bold', 25), bg='gray', command=main, text='PLAY', height=2, width=12)
play_but.grid(row=1, column=1)
label_1 = Label(root, font=('bold', 18), width=63, text='BUTTON A DESCRIPTION', height=0, bg='magenta', anchor='e')
label_1.grid(row=2, column=0)
button_a = Button(root, text='A', fg='magenta', font=('bold', 15), borderwidth=4, width=22, command=AAAA)
button_a.grid(row=2, column=1)
label_2 = Label(root, font=('bold', 18), width=63, text='BUTTON B DESCRIPTION', height=0, bg='green1', anchor='e')
label_2.grid(row=3, column=0)
button_b = Button(root, text='B', fg='green1', font=('bold', 15), borderwidth=4, width=22, command=BBBB)
button_b.grid(row=3, column=1)
label_3 = Label(root, font=('bold', 18), width=63, text='BUTTON C DESCRIPTION', height=0, bg='#7F7FFF', anchor='e')
label_3.grid(row=4, column=0)
button_c = Button(root, text='C', fg='#7F7FFF', font=('bold', 15), borderwidth=4, width=22, command=CCCC)
button_c.grid(row=4, column=1)
label_4 = Label(root, font=('bold', 18), width=63, text='  ENTER YOUR WORD HERE', height=0, bg='gray60', anchor='e')
label_4.grid(row=5, column=0)
guess = Entry(root, width=24, font=('bold', 16))
guess.bind('<Return>', check)
guess.grid(row=5, column=1)
defi_level = Button(root, text='difficulty', font=('bold', 18), height=1, width=12, borderwidth=3, command=difficulty)
defi_level.grid(row=2, column=2)
helps = Button(root, text='help', font=('bold', 18), height=1, width=12, borderwidth=3, command=helps_1)
helps.grid(row=3, column=2)
input_method = Button(root, text='input method', font=('bold', 18), height=1, width=12, borderwidth=3, command=input_1)
input_method.grid(row=4, column=2)
cDif = Label(root, text=show, font=('bold', 15))
cDif.grid(row=2, column=3)

root.mainloop()